import requests
from openpyxl import Workbook
import os
import json
from tencentcloud.common import credential
from tencentcloud.common.profile.client_profile import ClientProfile
from tencentcloud.common.profile.http_profile import HttpProfile
from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
from tencentcloud.ocr.v20181119 import ocr_client, models
import base64


def get_filename_from_dir(dir_path):
    file_list = []
    for item in os.listdir(dir_path):
        basename = os.path.basename(item)
        file_list.append(basename)

    return file_list


def getresult(file):
    data = []
    import ssl
    ssl._create_default_https_context = ssl._create_unverified_context
    try:
        cred = credential.Credential("AKIDhCSKyOceP5TUIFfh2Bdsrc7fzFybI2tV", "e3B0bx3siJdkaBauxFI6dugoeBUmsinf")
        httpProfile = HttpProfile()
        httpProfile.endpoint = "ocr.tencentcloudapi.com"

        clientProfile = ClientProfile()
        clientProfile.httpProfile = httpProfile
        client = ocr_client.OcrClient(cred, "ap-shanghai", clientProfile)

        req = models.VatInvoiceOCRRequest()
        with open("image/"+file, "rb") as f:
            filedate = base64.b64encode(f.read())
        params = {
            "ImageBase64": str(filedate,'utf-8')
        }
        req.from_json_string(json.dumps(params))

        resp = client.VatInvoiceOCR(req)
        result = resp.to_json_string()
        result = eval(result)
        for msg in result['VatInvoiceInfos']:
            if msg["Name"]=='打印发票号码':
                data.insert(0,msg["Value"])
            if msg["Name"]=='购买方名称':
                data.insert(1,msg["Value"])
            if msg["Name"]=='货物或应税劳务、服务名称':
                data.insert(2,msg["Value"])
            if msg["Name"]=='数量':
                data.insert(3,msg["Value"])
            if msg["Name"]=='小写金额':
                data.insert(4,msg["Value"])
            if msg["Name"]=='销售方名称':
                data.insert(5,msg["Value"])
            if msg["Name"]=='Name':
                data.insert(6,msg["Value"])



    except TencentCloudSDKException as err:
        print(err)

    return data

def write_excel(data0):
    wb = Workbook()
    # 写入表头
    dilei_head = ['发票号码', '购买方名称', '货物或应税劳务、服务名称', '数量', '价税合计', '销售方名称', '备注']
    sheet0Name = '发票信息'
    sheet0 = wb.create_sheet(sheet0Name, index=0)
    for i, item in enumerate(dilei_head):
        sheet0.cell(row=1, column=i + 1, value=item)
    # 写入数据
    for i, item in enumerate(data0):
        i = i + 1
        for j, val in enumerate(item):
            sheet0.cell(row=i, column=j + 1, value=val)

    wb.save("result" + '.xlsx')
    print("发票信息保存")
def main():
    file_list = get_filename_from_dir("image")
    datas = []
    for file in file_list:
        data = getresult(file)
        datas.append(data)

    write_excel(datas)


if __name__ == '__main__':
    main()


