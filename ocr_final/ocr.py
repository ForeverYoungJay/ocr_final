import requests
from openpyxl import Workbook
import os
import json
from tencentcloud.common import credential
from tencentcloud.common.profile.client_profile import ClientProfile
from tencentcloud.common.profile.http_profile import HttpProfile
from tencentcloud.common.exception.tencent_cloud_sdk_exception import TencentCloudSDKException
from tencentcloud.ocr.v20181119 import ocr_client, models
import base64
import fitz
import time


def pdf_image():
    for pdf in os.listdir("pdf"):
        pdf_name = "pdf/"+str(pdf)
    img_paths = []
    print("正在处理pdf:" + str(pdf))
    pdf = fitz.Document(pdf_name)
    for i,pg in enumerate(range(0, pdf.pageCount)):
        page = pdf[pg]  # 获得每一页的对象
        trans = fitz.Matrix(3.0, 3.0).preRotate(0)
        pm = page.getPixmap(matrix=trans, alpha=False)  # 获得每一页的流对象
        # pm.writePNG(dir_name + os.sep + base_name[:-4] + '_' + '{:0>3d}.png'.format(pg + 1))  # 保存图片
        img_path = str(pdf)+"_"+str(pg+1) + '.png'
        pm.writePNG("image/"+img_path)  # 保存图片
        img_paths.append(img_path)
    pdf.close()
    return img_paths

def get_filename_from_dir(dir_path):
    file_list = []
    for item in os.listdir(dir_path):
        basename = os.path.basename(item)
        file_list.append(basename)

    return file_list


def getresult(file):
    data = []
    import ssl
    ssl._create_default_https_context = ssl._create_unverified_context
    try:
        cred = credential.Credential("AKIDhCSKyOceP5TUIFfh2Bdsrc7fzFybI2tV", "e3B0bx3siJdkaBauxFI6dugoeBUmsinf")
        httpProfile = HttpProfile()
        httpProfile.endpoint = "ocr.tencentcloudapi.com"

        clientProfile = ClientProfile()
        clientProfile.httpProfile = httpProfile
        client = ocr_client.OcrClient(cred, "ap-shanghai", clientProfile)

        req = models.VatInvoiceOCRRequest()
        with open("image/"+file, "rb") as f:
            filedate = base64.b64encode(f.read())
        params = {
            "ImageBase64": str(filedate,'utf-8')
        }
        req.from_json_string(json.dumps(params))

        resp = client.VatInvoiceOCR(req)
        result = resp.to_json_string()
        result = json.loads(result)

        for msg in result['VatInvoiceInfos']:
            if msg["Name"]=='打印发票号码':
                num = msg["Value"][2:]
                if num[0]=="0":
                    num = "'"+num
                data.insert(0,num)
            if msg["Name"]=='销售方名称':
                data.insert(1,msg["Value"])
            if msg["Name"]=='备注':
                data.insert(2,msg["Value"])
            if msg["Name"]=='合计金额':
                data.insert(3,msg["Value"])
            if msg["Name"]=='合计税额':
                data.insert(4,msg["Value"])
            if msg["Name"]=='小写金额':
                if msg["Value"][-2:] == "00":
                    data.insert(5, msg["Value"][:-3])
                else:
                    data.insert(5,msg["Value"])
            if msg["Name"]=='购买方名称':
                name  = msg["Name"]
                if msg["Value"]=="上海晓篪信息技术有限公司":
                    name = "Xiaochi"
                elif msg["Value"]=="上海兢展电子科技有限公司":
                    name = "Jingzhan"
                elif msg["Value"]== "上海巢威计算机科技有限公司":
                    name = "Chaowei"
                elif msg["Value"]== "上海楷晟信息技术有限公司":
                    name = "Kaisheng"
                elif msg["Value"]== "上海翊瑄电子科技有限公司":
                    name = "Yixuan"

                data.insert(6,name)


        print("正在处理发票"+file)
        os.remove("image/"+file)
    except TencentCloudSDKException as err:
            print(err)
            print("未识别"+file)



    return data

def write_excel(data0):
    wb = Workbook()
    # 写入表头
    dilei_head = ['发票号码', '供应商名称', 'PO', '未税金额', '税额', '含税额', '公司名称']
    sheet0Name = '发票信息'
    sheet0 = wb.create_sheet(sheet0Name, index=0)
    for i, item in enumerate(dilei_head):
        sheet0.cell(row=1, column=i + 1, value=item)
    # 写入数据
    for i, item in enumerate(data0):
        i = i + 1
        for j, val in enumerate(item):
            sheet0.cell(row=i, column=j + 1, value=val)

    wb.save("result" + '.xlsx')
    print("发票信息保存成功")
def main():
    pdf_image()
    file_list = get_filename_from_dir("image")
    datas = []
    for file in file_list:
        data = getresult(file)
        datas.append(data)

    write_excel(datas)



if __name__ == '__main__':
    main()


